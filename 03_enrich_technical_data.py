import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os
import sys

# Function to print formatted log messages with timestamps and emojis
def print_log(message, emoji="INFO"):
    """
    Logs a message with a timestamp and optional emoji.
    Falls back to plain text if encoding issues arise.
    """
    try:
        emoji = emoji.encode("ascii", "ignore").decode("ascii") if not sys.stdout.encoding.startswith("utf") else emoji
    except UnicodeEncodeError:
        emoji = "INFO"  # Fallback if encoding fails

    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    print(f"[{timestamp}] {emoji} {message}")

# Function to read configuration file
def read_config(config_file):
    try:
        with open(config_file, 'r') as f:
            config = {}
            for line in f:
                key, value = line.strip().split('=')
                config[key.strip()] = value.strip()
        print_log(f"Configuration loaded successfully from {config_file}", "✅")
        return config
    except Exception as e:
        print_log(f"Error reading configuration file: {e}", "❌")
        raise

def main():
    config_file = 'merge.conf.txt'

    # Load configuration
    try:
        config = read_config(config_file)
    except Exception:
        return

    # Input file names from config
    technical_data_path = config.get('cdgc_file', 'cdgc_snowflake_table_export.xlsx')
    usage_stats_path = config.get('usage_file', 'usage_stats_snowflake.csv')

    print_log(f"Loading input files: {technical_data_path} and {usage_stats_path}")

    try:
        # Load the Excel file and the Technical Data Set sheet
        workbook = load_workbook(technical_data_path)
        sheet = workbook['Technical Data Set']
        technical_data_set = pd.DataFrame(sheet.values)
        technical_data_set.columns = technical_data_set.iloc[0]
        technical_data_set = technical_data_set[1:]  # Drop header row from values
        usage_stats = pd.read_csv(usage_stats_path)
        print_log("Input files loaded successfully", "✅")
    except Exception as e:
        print_log(f"Error loading input files: {e}", "❌")
        return

    # Prepare join keys
    try:
        usage_stats['join_key'] = usage_stats['DATABASE_NAME'] + '/' + usage_stats['SCHEMA_NAME'] + '/' + usage_stats['OBJECT_NAME']
        if (technical_data_set['HierarchicalPath'].str.len() > 2).all():
            technical_data_set['join_key'] = technical_data_set['HierarchicalPath'].str.split(pat='/', n=1).str[1]
        else:
            ## Oh no. The HierarchicalPath for some or all of these are blank!
            ## Using the Reference ID instead
            technical_data_set['join_key'] = technical_data_set['Reference ID'].str.extract(r'//(.*?)~')   

        ## print(f"DEBUGSCOTT: Setting join key on usage_stats: {usage_stats['join_key']}")
       
        # Ensure unique keys in usage_stats by grouping (e.g., taking the first occurrence)
        usage_stats_grouped = usage_stats.groupby('join_key').first().reset_index()

        print_log("Join keys prepared and duplicates handled successfully", "✅")
    except Exception as e:
        print_log(f"Error preparing join keys: {e}", "❌")
        return

    # Update values for specified columns
    try:
        # Create a lookup dictionary from the grouped usage_stats
        lookup = usage_stats_grouped.set_index('join_key')[['TOTAL_ACCESS_COUNT', 'ACCESS_COUNT_LAST_3_MONTHS', 'LAST_ACCESSED', 'LAST_UPDATED', 'LAST_QUERIES']].to_dict('index')

        # Update the corresponding columns in the technical data set
        for idx, row in technical_data_set.iterrows():
            key = row['join_key']
            if key in lookup:
                technical_data_set.at[idx, 'Total Access Count'] = lookup[key]['TOTAL_ACCESS_COUNT']
                technical_data_set.at[idx, 'Access Count Last 3 Months'] = lookup[key]['ACCESS_COUNT_LAST_3_MONTHS']
                technical_data_set.at[idx, 'Last Query Execution Date'] = lookup[key]['LAST_ACCESSED']
                technical_data_set.at[idx, 'Data Freshness'] = lookup[key]['LAST_UPDATED']
                technical_data_set.at[idx, 'Recent Queries'] = lookup[key]['LAST_QUERIES']

        print_log("Columns updated successfully", "✅")
    except Exception as e:
        print_log(f"Error updating columns: {e}", "❌")
        return

    # Drop the temporary join_key column
    technical_data_set.drop(columns=['join_key'], inplace=True)

    # Save the enriched data back to the Excel file while preserving other sheets
    enriched_file_path = os.path.splitext(technical_data_path)[0] + "_enriched.xlsx"

    try:
        # Update the "Technical Data Set" sheet with enriched data
        sheet.delete_rows(1, sheet.max_row)
        for r_idx, row in enumerate(dataframe_to_rows(technical_data_set, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

        # Save the workbook with all sheets intact
        workbook.save(enriched_file_path)
        print_log(f"Updated file saved as '{enriched_file_path}' successfully", "✅")
    except Exception as e:
        print_log(f"Error saving updated file: {e}", "❌")
        return

if __name__ == '__main__':
    main()
